"""Build the MH2 tracking dictionary from the supplied MH2 workbook."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


EVENT_COLUMNS = 9


def clean(value: object | None) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).replace("|", "\\|")


def markdown_table(headers: list[str], rows: list[list[str]]) -> str:
    lines = ["| " + " | ".join(headers) + " |"]
    lines.append("| " + " | ".join("---" for _ in headers) + " |")
    lines.extend("| " + " | ".join(row) + " |" for row in rows)
    return "\n".join(lines)


def read_rows(sheet) -> list[list[str]]:
    return [
        [clean(value) for value in row]
        for row in sheet.iter_rows(values_only=True)
        if any(value is not None for value in row)
    ]


def build(workbook_path: Path, output_path: Path) -> None:
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    identity_rows = read_rows(workbook["#用户ID体系"])
    common_rows = read_rows(workbook["#公共事件属性"])
    user_rows = read_rows(workbook["#用户数据"])

    event_sheet = workbook["#事件数据"]
    inherited_event = [""] * 5
    events: dict[str, list[str]] = {}
    properties: list[list[str]] = []

    for raw_row in event_sheet.iter_rows(min_row=2, max_col=EVENT_COLUMNS, values_only=True):
        if not any(value is not None for value in raw_row):
            continue
        event_columns = [clean(value) for value in raw_row[:5]]
        inherited_event = [
            value if value else inherited_event[index]
            for index, value in enumerate(event_columns)
        ]
        event_name = inherited_event[0]
        if not event_name:
            continue
        events.setdefault(event_name, inherited_event[1:])
        property_name = clean(raw_row[5])
        if property_name:
            properties.append(
                [event_name, *inherited_event[1:], property_name, clean(raw_row[6]), clean(raw_row[7]), clean(raw_row[8])]
            )

    lines = [
        "# MH2 埋点词典",
        "",
        f"来源：`{workbook_path.name}`。本文件由 `scripts/build_tracking_dictionary.py` 生成。",
        "",
        "字段的技术名、中文含义、类型和枚举以该表为准。字段不在这里时，不能据此说它不存在；应向用户索要服务端表、配置表或另一份数数 Schema 来源。",
        "",
        "## 用户 ID 体系",
        "",
        markdown_table(identity_rows[0][:5], [row[:5] for row in identity_rows[1:]]),
        "",
        "## 公共事件属性",
        "",
        markdown_table(common_rows[0][:4], [row[:4] for row in common_rows[1:]]),
        "",
        "## 用户属性",
        "",
        markdown_table(user_rows[0][:6], [row[:6] for row in user_rows[1:]]),
        "",
        "## 事件",
        "",
        markdown_table(
            ["事件名", "事件中文名", "说明", "来源端", "分类"],
            [[name, *metadata] for name, metadata in events.items()],
        ),
        "",
        "## 事件属性",
        "",
        markdown_table(
            ["事件名", "事件中文名", "说明", "来源端", "分类", "属性名", "属性中文名", "类型", "说明"],
            properties,
        ),
        "",
    ]
    output_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    build(args.workbook, args.output)


if __name__ == "__main__":
    main()
